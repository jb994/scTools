import numpy as np
import pandas as pd
import scanpy as sc
import matplotlib as mpl
import matplotlib.pyplot as plt
from scTools import process
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

import warnings
from anndata import ImplicitModificationWarning
warnings.filterwarnings("ignore", category=ImplicitModificationWarning)

def makeSpreadsheet(df, filename=None, sheetKey = 'cellType', sortKey=None, columnDict=None, index=False, speak=True):
	### Given a df and a sheetKey, will create an excel spreadsheet where each sheet\
	### is for each unique value in the sheetKey column
	if filename is None:
		print("Must give valid filename location")
		quit()
	wb = Workbook()
	ws1 = wb.active
	ws1.title = "All"
	for r in dataframe_to_rows(df, index=index, header=True):
		ws1.append(r)

	for key in np.unique(df[sheetKey]):
		### Makes each sheet
		if speak: print(f"On {key}")
		keyDF = df[df[sheetKey]==key]
		if sortKey: keyDF = keyDF.sort_values(sortKey)

		if columnDict is None:
			columns=df.columns
		else:
			columns = columnDict[key]

		ws = wb.create_sheet(title=key)
		for r in dataframe_to_rows(keyDF[columns], index=index, header=True):
			ws.append(r)

	wb.save(filename)
	print(f"File {filename} created")


def picsToSpreadsheet(spreadsheetPath, images, titles=["Sheet1"]):
	### images is a list of imagePaths or a list of lists for each sheet
	### titles is a list of sheetTitles
	wb=Workbook()
	#ws=wb.active

	for sheetIndex, title in enumerate(titles):
		ws = wb.create_sheet(title=title)
		if len(titles)==1: images = [images]
		for i, image in enumerate(images[sheetIndex]):
			img = Image(image)
			ws.add_image(img, f'B{12*i+1}')

	wb.save(spreadsheetPath)

def rankCellTypes(Data, treatments, silent=False, heatmap=True, norm=True, key='fineClusters'):
	### Put in a dictionary of AnnDatas and their keys
	### For each pair of AnnDatas a correlation score is determined for\
	### how much each of the celltypes have changed between groups.
	### Returns a DF of treatment comparisons and cellTypes
	cols=[]
	dfData=[]
	allTypes = np.unique(process.catAdata(Data, treatments).obs[key])

	for i,t1 in enumerate(treatments):
		for j,t2 in enumerate(treatments):
			if i<j:
				print(f"{t1} x {t2}")
				cols.append(f"{t1} x {t2}")
				bigData = process.catAdata(Data, [t1,t2])
				if norm:
					bigData = process.dgeNorm(bigData, useRaw=False, scale=False, pca=True)
				row=[]
				for cellType in allTypes:
					if not silent: print(cellType)
					if process.checkClusterInAdatas([bigData[bigData.obs['batch']==t1],bigData[bigData.obs['batch']==t2]], cellType, obsKey=key):
					#if (cellType in np.unique(bigData[bigData.obs['batch']==t1].obs[key])) and (cellType in np.unique(bigData[bigData.obs['batch']==t2].obs[key])):
						adata = bigData[bigData.obs[key]==cellType].copy()
						sc.pp.scale(adata)
						sc.tl.rank_genes_groups(adata, 'batch', method='wilcoxon', use_raw=False, max_iter=2000, pts=False,cor_method='spearman')
						sc.tl.dendrogram(adata, groupby='batch', )
						if not silent: print(adata.uns['dendrogram_batch']['correlation_matrix'][0,1])
						if heatmap: sc.pl.rank_genes_groups_heatmap(adata, vmin=-3, vmax=3, cmap='bwr')
						# mpl.colors.SymLogNorm or mpl.colors.CenteredNorm #norm=mpl.colors.CenteredNorm()
						row.append(adata.uns['dendrogram_batch']['correlation_matrix'][0,1])
					else:
						row.append(None)

				dfData.append(row)
	return pd.DataFrame(np.asarray(dfData).T, columns=cols, index=allTypes)




